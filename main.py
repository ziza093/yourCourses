import urllib.request
import requests
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter
from copy import copy

FORMATS = ["C s", "C p", "C i", "P s", "P p", "P i", "S s", "S p", "S i", "L s", "L i", "L p"]

def get_file():
    #this is the website from which I download the file (spreadsheet)
    site_url = "https://ac.tuiasi.ro/studenti/didactic/orar/"

    #the document link, but calculated later
    doc_url = ""

    #try to reach the website
    site_response = urllib.request.urlopen(site_url)
    #extract its content
    site_content = site_response.read()

    #this is the phrase the help me extract the spreadsheet
    phrase = "orar ac"

    #iterate through every line of code from the website untill I reach the line which has my spreadsheet link and extract the link
    for line in site_content.decode("utf-8").splitlines():
        if phrase in line.lower():
            doc_url = line.split("\"")

            for elem in doc_url:
                if "https" in elem:
                    doc_url = elem
                    break
            break

    #extract from the url the spreadsheet id and complete it with the file format extension to export it
    doc_url = doc_url.split("/edit")[0]+"/export?format=xlsx"

    #make a get request to the google sheet
    response = requests.get(doc_url)

    #write its contents to an local file
    with open("orar_full.xlsx", "wb") as f:
        f.write(response.content)


def get_group_col(ws_source):

    group_col = None
    while group_col is None:
        try:
            print(f"Choose an group/class")
            group_name = input()

            for ro in range(1, ws_source.max_row):
                for col in range(1, ws_source.max_column):
                    source_cell = ws_source.cell(row=ro, column=col)
                    source_cell_value = str(source_cell.value).lower()

                    if source_cell_value == group_name.lower():
                        group_col = source_cell.column

        except ValueError("Group not in the sheet!"):
            print(f"Try again!")
            group_name = input()
    
    return group_col



def parse_format(index, name):     
    #extract all the words from the course/project/seminar
    parts = name.split(FORMATS[index])
    parts[0] = parts[0].split()
    parts[1] = parts[1].split()    

    #the room is the last word of the second list
    room = parts[1][-1]
    #the course name is the last word of the first list
    course_name = parts[0][-1]
    # #get the word with the course name in the paranthesis
    # for part in parts:
    #     if '(' in part:
    #         course_name = part
    #         break
    
    #go through word and extract the one before the frequency
    # print(f"parts: {parts}")
    # if course_name == "":
    #     for i in range(0, len(parts)):
    #         # if FORMATS[index] parts[i]
    #         pass
    
    #remove the paranthesis (BD)
    if course_name[0] == '(' and course_name[-1] == ')':
        course_name = course_name[1:]
        course_name = course_name[:len(course_name)-1]

    frequency = FORMATS[index]
    frequency[2].lower()

    new_name = f"{course_name.upper()} {frequency}\n{room.upper()}"

    return new_name


def get_courses(ws_source, group_col):

    courses_list = {}

    for ro in range(1, ws_source.max_row):
        for col in range(1, group_col + 1):
            source_cell = ws_source.cell(row=ro, column=col)
            source_cell_value = str(source_cell.value)
            
            for form in FORMATS:
                if form in source_cell_value:
                    #check if this cell is merged and if the merge includes group_col
                    cell_coord = source_cell.coordinate
                    is_in_merged_range = False
                    
                    for merged_range in ws_source.merged_cells.ranges:
                        if cell_coord in merged_range:
                            #check if group_col is within this merged range
                            if merged_range.min_col <= group_col <= merged_range.max_col:
                                is_in_merged_range = True
                            break
                    
                    # Add only if: merged and includes group_col, OR not merged and is group_col
                    if is_in_merged_range or col == group_col:
                        courses_list.update({
                            source_cell_value: {
                                'row': ro, 
                                'fill': copy(source_cell.fill)
                            }
                        })     
                        #dont duplicate the same course          
                        break      


    #modify the courses_list so that i have the final form of the courses string
    modified_courses = {}
    for course in courses_list:
        for index in range(0, len(FORMATS)):
            if FORMATS[index] in course and course not in modified_courses:
                parsed_string = parse_format(index, course)
                modified_courses.update({parsed_string:courses_list[course]})

    return modified_courses


def get_cells(ws_source, group_col):
    cells_list = {}

    for ro in range(1, ws_source.max_row):
        source_cell = ws_source.cell(row=ro, column=group_col)
        source_cell_value = str(source_cell.value)
        if source_cell_value:
            cells_list.update({
                source_cell_value: {
                    'row': ro, 
                    'fill': copy(source_cell.fill)
                }
            })

    #modify the cells_list so that i have the final form of the courses string
    modified_cell = {}
    for cell in cells_list:
        for index in range(0, len(FORMATS)):
            if FORMATS[index] in cell and cell not in modified_cell:
                parsed_string = parse_format(index, cell)
                modified_cell.update({parsed_string:cells_list[cell]})

    return modified_cell

def get_weekdays(ws_source):
    weekdays = {
    "luni" : [],
    "marți" : [],
    "marti" : [],
    "miercuri" : [],
    "joi" : [],
    "vineri": []
    }

    for ro in range(1, ws_source.max_row):
        for col in range(1, ws_source.max_column):
            source_cell = ws_source.cell(row=ro, column=col)
            source_cell_value = str(source_cell.value).lower()

            if source_cell_value in weekdays:
                weekdays[source_cell_value] = ro

    return weekdays


def set_format_cells(ws_personal):

    for ro in range(1, ws_personal.max_row+1):
        for col in range(2, ws_personal.max_column+1):
            personal_cell = ws_personal.cell(row=ro, column=col)
            personal_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')



def extract_table():
    
    wb_general = openpyxl.load_workbook(filename="orar_full.xlsx")

    sheet_name = ""

    print(f"Choose an year and specialization from the list: {wb_general.sheetnames}")
    sheet_name = input()

    while(sheet_name not in wb_general.sheetnames):
        print("Doesnt fit any of the ones in the list! Please write one from the list!")
        print(wb_general.sheetnames)
        sheet_name = input()


    ws_source = wb_general.worksheets[wb_general.sheetnames.index(sheet_name)]
    

 

    #get the column pozition in which your group is situated
    group_col = get_group_col(ws_source)

    #extract only the courses the group attends
    courses_list = get_courses(ws_source, group_col)


    #extract all the cells 
    cells_list = get_cells(ws_source, group_col)

    #have the choice to delete unwanted courses from there
    remove_unwanted_cells(cells_list, courses_list)

    #extract the weekdays positions
    weekdays = get_weekdays(ws_source)

    #create the workbook
    wb_personal = create_table(courses_list, cells_list, weekdays)

    #save to file
    wb_personal.save("table.xlsx")


def add_personal_all_data(ws_personal, weekday, courses_list, cells_list, weekdays, week_col):
    # Helper function to write data and style
    def write_cell(target_row, target_col, text, fill_obj):
        cell = ws_personal.cell(row=target_row, column=target_col)
        cell.value = text
        cell.fill = fill_obj

    # --- Process Courses ---
    for course_name, data in courses_list.items():
        course_row = int(data['row']) 
        weekday_row = int(weekdays[weekday])
        
        if course_row >= weekday_row and course_row <= weekday_row + 11:
            target_row = 0
            if course_row == weekday_row:
                target_row = 2
            else:
                target_row = 2 + course_row - weekday_row
            
            # Write value AND color
            write_cell(target_row, week_col, course_name, data['fill'])

    # --- Process Labs/Seminars (Cells) ---
    for cell_name, data in cells_list.items():
        cell_row = int(data['row'])
        weekday_row = int(weekdays[weekday])

        if cell_row >= weekday_row and cell_row <= weekday_row + 11:
            target_row = 0
            if cell_row == weekday_row:
                target_row = 2
            else:
                target_row = 2 + cell_row - weekday_row
            
            # Write value AND color
            write_cell(target_row, week_col, cell_name, data['fill'])


def merge_final_cells(ws_personal):
    for col in range(2, ws_personal.max_column + 1):
        for ro in range(2,ws_personal.max_row):
            personal_cell = ws_personal.cell(row=ro, column=col)
            next_personal_cell = ws_personal.cell(row=ro+1, column=col)
            
            if personal_cell.value:
                if not next_personal_cell.value:
                    if "p i" in personal_cell.value or "p p" in personal_cell.value:
                        ro = ro + 2
                    else:
                        ws_personal.merge_cells(start_row=ro, start_column=col, end_row=ro+1, end_column=col)
                        ro = ro + 1


def remove_unwanted_cells(cells_list, courses_list):

    print("Do you want to remove some of the contents of your table?")
    print("Type YES if so, otherwise type NO")
    if input() == "NO":
        return

    print("If you want to save the rest of the remaining cells, quit by writing 'EXIT'")
    print("If you wanna keep going, write 'CONTINUE'")

    choice = input()

    while choice == "CONTINUE":
        keys_cells = list(cells_list.keys())
        keys_courses = list(courses_list.keys())

        total_items = []

        i=0
        print("\nLabs/Seminars\n")
        for key in cells_list:
            i = i+1
            print(f"{i}. {key}")
            total_items.append(('cell', key))

        print("\nCourses\n")
        for key in courses_list:
            i = i+1
            print(f"{i}. {key}")
            total_items.append(('course', key))
            
        print("Select which number to delete!")
        try:
            j = int(input())
            if 1 <= j <= len(total_items):
                type_to_del, key_to_del = total_items[j-1]

                if type_to_del == 'cell':
                    cells_list.pop(key_to_del)
                else:
                    courses_list.pop(key_to_del)
                print(f"Deleted: {key_to_del}")
            else:
                print("Invalid number!") 
        except ValueError:
            print("Please enter a valid number!")
        
        print("If you want to save the rest of the remaining cells, quit by writing 'EXIT'")
        print("If you wanna keep going, write 'CONTINUE'")
        choice = input()


def create_table(courses_list, cells_list, weekdays):
    #create the workbook
    wb_personal = openpyxl.Workbook()

    #select the only sheet in the workbook
    ws_personal = wb_personal.active
    
    ws_personal.cell(row=1, column=2).value = "LUNI"
    ws_personal.cell(row=1, column=3).value = "MARTI"
    ws_personal.cell(row=1, column=4).value = "MIERCURI"
    ws_personal.cell(row=1, column=5).value = "JOI"
    ws_personal.cell(row=1, column=6).value = "VINERI"

    #weekdays
    for i in range(1,7):
        ws_personal.cell(row=1, column=i).font = openpyxl.styles.Font(bold=True)

    #hours every day
    j = 8
    for i in range(2, 14):
        ws_personal.cell(row=i, column=1).value = j
        j = j + 1


    for i in range(1,14):
        for j in range(1, 7):
            ws_personal.cell(row=i, column=j).border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style="thin", color = "000000"),
                                                                            right=openpyxl.styles.Side(border_style="thin", color= "000000"),
                                                                            bottom=openpyxl.styles.Side(border_style="thin", color="000000"),
                                                                            top=openpyxl.styles.Side(border_style="thin", color="000000"))

    
    for i in range(1, 7):
        #first column 
        if i == 1:
            size = 1
        else:
            size = len(str(ws_personal.cell(row=1, column=i).value))
        
        letter = get_column_letter(i)
        ws_personal.column_dimensions[letter].width = (size + 2) * 1.5

    
    add_personal_all_data(ws_personal, "luni", courses_list, cells_list, weekdays, 2)
    try:
        add_personal_all_data(ws_personal, "marți", courses_list, cells_list, weekdays, 3)
    except:
        add_personal_all_data(ws_personal, "marti", courses_list, cells_list, weekdays, 3)

    add_personal_all_data(ws_personal, "miercuri", courses_list, cells_list, weekdays, 4)
    add_personal_all_data(ws_personal, "joi", courses_list, cells_list, weekdays, 5)
    add_personal_all_data(ws_personal, "vineri", courses_list, cells_list, weekdays, 6)

    merge_final_cells(ws_personal)
    merge_final_cells(ws_personal)
    merge_final_cells(ws_personal)
    merge_final_cells(ws_personal)
    merge_final_cells(ws_personal)

    set_format_cells(ws_personal)

    return wb_personal

def main():
    get_file()
    extract_table()

if __name__ == "__main__":
    main()
